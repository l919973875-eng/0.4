from __future__ import annotations

import argparse
import csv
import hashlib
import json
import re
from datetime import datetime, timedelta, timezone
from pathlib import Path


def iso_time(value):
    if value in (None, ''):
        return None
    try:
        if isinstance(value, (int, float)) or str(value).isdigit():
            n = float(value)
            if n > 1e12:
                n /= 1000
            return datetime.fromtimestamp(n, tz=timezone.utc).isoformat()
    except Exception:
        pass
    try:
        from dateutil import parser as dtparser
        dt = dtparser.parse(str(value))
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=timezone.utc)
        return dt.astimezone(timezone.utc).isoformat()
    except Exception:
        return None


def first(d, *keys):
    for k in keys:
        v = d.get(k)
        if v not in (None, '', [], {}):
            return v
    return None


def to_int(v):
    try:
        if isinstance(v, str):
            v = re.sub(r'[^0-9.-]', '', v)
        return int(float(v or 0))
    except Exception:
        return 0


def normalize(platform: str, d: dict) -> dict | None:
    title = first(d, 'title', 'note_title', 'video_title') or ''
    desc = first(d, 'desc', 'content', 'text', 'note_desc', 'aweme_desc', 'description') or ''
    text = re.sub(r'\s+', ' ', f'{title} {desc}').strip()
    if not text:
        return None
    author = first(d, 'nickname', 'user_nickname', 'author_name', 'screen_name', 'user_name', 'author') or 'unknown'
    raw_id = first(d, 'note_id', 'aweme_id', 'id', 'mblogid', 'content_id', 'video_id') or hashlib.sha256(text.encode('utf-8')).hexdigest()[:20]
    url = first(d, 'note_url', 'aweme_url', 'video_url', 'url', 'content_url') or ''
    if not url:
        if platform == 'xiaohongshu':
            url = f'https://www.xiaohongshu.com/explore/{raw_id}'
        elif platform == 'douyin':
            url = f'https://www.douyin.com/video/{raw_id}'
        elif platform == 'weibo':
            url = f'https://weibo.com/detail/{raw_id}'
    published = iso_time(first(d, 'time', 'create_time', 'created_at', 'publish_time', 'note_time', 'last_modify_ts'))
    engagement = {
        'like_count': to_int(first(d, 'liked_count', 'like_count', 'attitudes_count', 'digg_count')),
        'comment_count': to_int(first(d, 'comment_count', 'comments_count')),
        'share_count': to_int(first(d, 'share_count', 'reposts_count')),
    }
    media = first(d, 'image_list', 'images', 'video_cover', 'video_url')
    media_count = len(media) if isinstance(media, list) else (1 if media else 0)
    sid = hashlib.sha256(f'{platform}:{raw_id}'.encode('utf-8')).hexdigest()[:20]
    return {
        'id': sid, 'platform': platform, 'author': str(author), 'text': text[:1800], 'url': str(url),
        'published_at': published, 'collected_at': datetime.now(timezone.utc).isoformat(), 'query': 'mediacrawler-search',
        'engagement': engagement, 'media_count': media_count, 'collector': 'MediaCrawler-external',
    }


def iter_rows(path: Path):
    suffix = path.suffix.lower()
    if suffix == '.json':
        obj = json.loads(path.read_text(encoding='utf-8'))
        rows = obj if isinstance(obj, list) else [obj]
        for row in rows:
            if isinstance(row, dict):
                yield row
    elif suffix == '.jsonl':
        for line in path.read_text(encoding='utf-8').splitlines():
            line = line.strip()
            if not line:
                continue
            try:
                row = json.loads(line)
                if isinstance(row, dict):
                    yield row
            except Exception:
                continue
    elif suffix == '.csv':
        with path.open('r', encoding='utf-8-sig', errors='replace', newline='') as f:
            yield from csv.DictReader(f)
    elif suffix == '.xlsx':
        try:
            from openpyxl import load_workbook
            wb = load_workbook(path, read_only=True, data_only=True)
            ws = wb['Contents'] if 'Contents' in wb.sheetnames else wb[wb.sheetnames[0]]
            rows = ws.iter_rows(values_only=True)
            headers = [str(x or '').strip() for x in next(rows)]
            for values in rows:
                yield dict(zip(headers, values))
        except Exception:
            return


def content_files(repo: Path, code: str) -> list[Path]:
    base = repo / 'data' / code
    if not base.exists():
        return []
    files = []
    for p in base.rglob('*'):
        if not p.is_file() or p.suffix.lower() not in {'.json', '.jsonl', '.csv', '.xlsx'}:
            continue
        low = p.name.lower()
        if any(x in low for x in ('comment', 'creator', 'wordcloud')):
            continue
        # 优先内容文件，同时兼容上游命名变化。
        if any(x in low for x in ('content', 'aweme', 'note', 'search')) or p.parent.name.lower() in {'json', 'jsonl', 'csv'}:
            files.append(p)
    return sorted(set(files))


def read_status_files(status_dir: Path) -> dict[str, dict]:
    out = {}
    for code in ('xhs', 'dy', 'wb'):
        p = status_dir / f'mediacrawler_{code}_status.json'
        if p.exists():
            try:
                obj = json.loads(p.read_text(encoding='utf-8'))
                if isinstance(obj, dict):
                    out[code] = obj
            except Exception:
                pass
    return out


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--repo', required=True)
    ap.add_argument('--output', required=True)
    ap.add_argument('--status-output', default='')
    ap.add_argument('--status-dir', default='data')
    args = ap.parse_args()
    repo = Path(args.repo)
    out_path = Path(args.output)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    mapping = {'xhs': 'xiaohongshu', 'dy': 'douyin', 'wb': 'weibo'}
    diagnostics = read_status_files(Path(args.status_dir))
    new = []
    status = []

    for code, platform in mapping.items():
        files = content_files(repo, code)
        count = 0
        parse_errors = 0
        for path in files:
            try:
                for d in iter_rows(path):
                    row = normalize(platform, d)
                    if row:
                        new.append(row)
                        count += 1
            except Exception:
                parse_errors += 1
        diag = diagnostics.get(code, {})
        if count:
            st = 'ok'
            detail = f'导入 {count} 条；读取 {len(files)} 个内容文件'
        elif diag:
            st = diag.get('status') or 'empty'
            detail = diag.get('detail') or '上游未产生内容'
        else:
            st = 'empty'
            detail = '未发现 MediaCrawler 内容文件；可能未配置 Cookie、平台返回空或采集步骤失败'
        status.append({
            'platform': platform, 'platform_code': code, 'status': st, 'items': count,
            'detail': detail, 'reason_code': diag.get('reason_code', ''), 'attempts': diag.get('attempts'),
            'files_found': len(files), 'parse_errors': parse_errors, 'collector': 'MediaCrawler',
        })

    try:
        existing = json.loads(out_path.read_text(encoding='utf-8')) if out_path.exists() else []
        if not isinstance(existing, list):
            existing = []
    except Exception:
        existing = []
    by_id = {r.get('id'): r for r in existing if r.get('id')}
    for r in new:
        by_id[r['id']] = r
    cutoff = datetime.now(timezone.utc) - timedelta(days=14)
    kept = []
    for r in by_id.values():
        try:
            dt = datetime.fromisoformat((r.get('published_at') or r.get('collected_at')).replace('Z', '+00:00'))
            if dt.tzinfo is None:
                dt = dt.replace(tzinfo=timezone.utc)
            if dt < cutoff:
                continue
        except Exception:
            pass
        kept.append(r)
    kept.sort(key=lambda x: x.get('published_at') or x.get('collected_at') or '', reverse=True)
    out_path.write_text(json.dumps(kept[:12000], ensure_ascii=False, indent=2), encoding='utf-8')
    if args.status_output:
        Path(args.status_output).write_text(json.dumps(status, ensure_ascii=False, indent=2), encoding='utf-8')
    print(f'ingested MediaCrawler: parsed={len(new)}, stored={len(kept)}')
    for s in status:
        print(f"  {s['platform']}: {s['status']} items={s['items']} reason={s.get('reason_code') or '-'} detail={s['detail']}")


if __name__ == '__main__':
    main()
